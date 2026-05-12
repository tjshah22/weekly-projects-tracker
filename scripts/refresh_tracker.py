#!/usr/bin/env python3
"""
Build a weekly leadership project tracker from a fulfillment update deck,
Jira, or both.

The app data refresh path intentionally uses only the standard library so it can
run in a locked-down corporate environment. PPTX files are ZIP packages with XML
inside, so the deck extraction is dependency-light and easy to audit.
"""

from __future__ import annotations

import argparse
import base64
import csv
import datetime as dt
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from xml.etree import ElementTree as ET
from zipfile import ZipFile

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    Workbook = None
    Alignment = Border = DataValidation = Font = PatternFill = Side = Table = TableStyleInfo = None
    get_column_letter = None


PML_NS = "http://schemas.openxmlformats.org/presentationml/2006/main"
DML_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"p": PML_NS, "a": DML_NS, "r": REL_NS, "rel": PKG_REL_NS}

STATUS_ALIASES = {
    "complete": "Completed",
    "completed": "Completed",
    "on track": "On Track",
    "on-track": "On Track",
    "monitoring": "Monitoring",
    "at risk": "At Risk",
    "blocked": "Blocked",
    "upcoming": "Upcoming",
}
STATUS_OPTIONS = ["Completed", "On Track", "Monitoring", "At Risk", "Blocked", "Upcoming"]
WORKSTREAM_OPTIONS = [
    "OTTO/Media Broker",
    "Fulfillment Modernization",
    "TLVOD/Fulfillment Pipeline",
    "CVP",
    "Business",
    "Department Metrics",
]

PROJECT_HEADERS = [
    "Reporting Week",
    "Workstream",
    "Project",
    "Status",
    "Owner",
    "Target Date",
    "This Week Update",
    "Next Milestone",
    "Business Impact",
    "Blocker or Risk",
    "Leadership Ask",
    "Help Needed",
    "Jira Key/Epic",
    "Jira URL",
    "Source Slide",
    "Source Deck",
    "Last Updated",
]

BLOCKER_HEADERS = [
    "Reporting Week",
    "Workstream",
    "Project",
    "Status",
    "Blocker or Risk",
    "Impact",
    "Owner",
    "Help Needed",
    "Target Date",
    "Source Slide",
]

HIGHLIGHT_HEADERS = [
    "Reporting Week",
    "Category",
    "Highlight",
    "Business Context",
    "Impact/Volume",
    "Timing",
    "Leadership Note",
    "Source Slide",
    "Source Deck",
]

METRIC_HEADERS = [
    "Reporting Period",
    "Metric Area",
    "Metric",
    "Value",
    "Unit/Breakdown",
    "Leadership Context",
    "Source Slide",
    "Source Deck",
]

JIRA_HEADERS = [
    "Run ID",
    "Source",
    "Query/Board",
    "Issue Key",
    "Issue Type",
    "Summary",
    "Status",
    "Status Category",
    "Assignee",
    "Priority",
    "Labels",
    "Components",
    "Fix Versions",
    "Due Date",
    "Updated",
    "Sprint",
    "Epic Key",
    "Epic Name",
    "Story Points",
    "URL",
]

SOURCE_HEADERS = ["Slide", "Slide Type", "Extracted Text", "Source Deck"]

BODY_STARTERS = (
    "real-time",
    "dev work",
    "tracking",
    "fast follow",
    "currently",
    "the cloud",
    "we need",
    "golf reporting",
    "workaround",
    "build ",
    "uncovered",
    "working with",
    "sending",
    "verified",
    "extend ",
    "metadata package",
    "estimated ",
    "investigation",
    "changes have",
    "auth window",
    "need ",
    "as part",
    "tmt to",
    "long from",
    "baton-sdvi",
    "temporary",
    "architecture",
    "contracts",
    "feasibility",
    "signiant",
    "rich and",
    "request for completion",
)

BUSINESS_HEADINGS = {
    "Future USA Sports Short Form Clipping",
    "Golf Channel Live Events In May",
    "Golf Channel Live Events In April",
    "USA Sports NASCAR Coverage in May",
    "WNBA Basketball Regular Season Games",
    "TLVOD-VTM integration",
    "Adobe Premiere Support Sessions",
    "Modernization Stakeholders",
    "ENT Archive Subclipping Screeners",
    "Smart Device Daily Briefing Skills",
    "Major Upcoming Live Events & Coverage USA Sports (2026+)",
}


def clean_text(value: str) -> str:
    replacements = {
        "\xa0": " ",
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    value = re.sub(r"\s+", " ", value)
    fixups = {
        "packa ge": "package",
        "ste ps": "steps",
        "thi s": "this",
        "wednesday": "Wednesday",
        "softops": "SoftOps",
        "ketki": "Ketki",
        "ivi": "IVI",
    }
    for old, new in fixups.items():
        value = re.sub(rf"\b{re.escape(old)}\b", new, value, flags=re.I)
    return value.strip()


def normalize_status(value: str) -> Optional[str]:
    key = clean_text(value).lower()
    return STATUS_ALIASES.get(key)


def is_footer_line(value: str) -> bool:
    cleaned = clean_text(value)
    if not cleaned:
        return True
    if cleaned.isdigit():
        return True
    if cleaned == "Presentation Name v1.0":
        return True
    if cleaned in {"Team", "Department Overview"}:
        return True
    if cleaned in {"Completed", "On Track", "Monitoring", "At Risk", "Blocked"}:
        return False
    return False


def slide_type(lines: Sequence[str]) -> str:
    joined = " | ".join(lines)
    if "Project Highlights" in joined:
        return "Project Highlights"
    if "Business Highlights" in joined:
        return "Business Highlights"
    if "Department Highlights" in joined:
        return "Department Highlights"
    if "Appendix" in joined:
        return "Appendix"
    if "Content Supply Chain" in joined:
        return "Title"
    return "Other"


def pptx_slide_paths(zip_file: ZipFile) -> List[str]:
    pres = ET.fromstring(zip_file.read("ppt/presentation.xml"))
    rels = ET.fromstring(zip_file.read("ppt/_rels/presentation.xml.rels"))
    rid_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    paths: List[str] = []
    slide_ids = pres.find("p:sldIdLst", NS)
    if slide_ids is None:
        return paths
    for slide_id in slide_ids:
        rid = slide_id.attrib[f"{{{REL_NS}}}id"]
        target = rid_to_target[rid]
        paths.append("ppt/" + target.lstrip("/"))
    return paths


def extract_slides(deck_path: Path) -> List[Dict[str, Any]]:
    slides: List[Dict[str, Any]] = []
    with ZipFile(deck_path) as zip_file:
        for slide_number, slide_path in enumerate(pptx_slide_paths(zip_file), start=1):
            root = ET.fromstring(zip_file.read(slide_path))
            lines: List[str] = []
            for text in root.findall(".//a:t", NS):
                cleaned = clean_text(text.text or "")
                if cleaned:
                    lines.append(cleaned)
            slides.append(
                {
                    "slide": slide_number,
                    "path": slide_path,
                    "lines": lines,
                    "text": "\n".join(lines),
                    "type": slide_type(lines),
                }
            )
    return slides


def likely_body_line(line: str, title_lines: Sequence[str]) -> bool:
    if not title_lines:
        return False
    lower = line.lower()
    if lower.startswith(BODY_STARTERS):
        return True
    title = " ".join(title_lines)
    if "|" in title and re.search(r"(\d{1,2}/\d{1,2}|TBD|Week of)", title, re.I):
        return not line.startswith("|")
    if len(title_lines) >= 3:
        return True
    return False


def should_join_title_fragment(fragment: str, title_lines: Sequence[str]) -> bool:
    if not title_lines:
        return True
    title = " ".join(title_lines)
    if title.endswith(("(", "{", "-", "|")):
        return True
    if fragment.startswith(("/", "|", "-")):
        return True
    if fragment.endswith(("}", ")")):
        return True
    if re.fullmatch(r"[A-Za-z0-9 /&.-]{1,24}", fragment) and not likely_body_line(fragment, title_lines):
        return True
    return False


def clean_project_title(value: str) -> str:
    value = clean_text(value)
    value = re.sub(r"(\d{1,2})\s*/\s*(\d{1,2})", r"\1/\2", value)
    value = value.replace("{", "(").replace("}", ")")
    value = re.sub(r"\(\s+", "(", value)
    value = re.sub(r"\s+\)", ")", value)
    value = value.replace("Build|", "Build |")
    value = re.sub(r"\s{2,}", " ", value)
    return value.strip()


def infer_workstream(project: str) -> str:
    p = project.lower()
    if any(token in p for token in ["polaris", "modernization", "auto qc", "content hub", "sdvi"]):
        return "Fulfillment Modernization"
    if p.startswith("cvp") or "wildmoka" in p or "viewlift" in p:
        return "CVP"
    if p.startswith("tlvod") or p.startswith("fp ") or "fulfillment pipeline" in p or "vtm" in p:
        return "TLVOD/Fulfillment Pipeline"
    if any(token in p for token in ["otto", "media broker", "dataforge", "rundown", "datadog"]):
        return "OTTO/Media Broker"
    return "Business"


def extract_owner(project: str) -> str:
    parts = [part.strip() for part in project.split("|")]
    if len(parts) >= 2:
        candidate = parts[-1]
        if candidate and not re.search(r"\d|TBD|Week", candidate, re.I):
            return candidate
    match = re.search(r"\|\s*([A-Z][A-Za-z]+(?:\s+[A-Z])?)\s*$", project)
    return match.group(1) if match else ""


def extract_target_date(text: str) -> str:
    patterns = [
        r"Week of \d{1,2}/\d{1,2}",
        r"\b\d{1,2}/\d{1,2}/\d{2,4}\b",
        r"\b\d{1,2}/\d{1,2}\b",
        r"\bTBD\b",
        r"\bQ[1-4]\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            return match.group(0)
    completion = re.search(r"(Estimated Time of Completion|Estimated time of completion|Request for Completion):?\s*([^.;\n]+)", text)
    return clean_text(completion.group(2)) if completion else ""


def next_milestone(project: str, body: str) -> str:
    combined = f"{project}. {body}"
    dates = re.findall(
        r"(?:\d{1,2}/\d{1,2}(?:/\d{2,4})?|Week of \d{1,2}/\d{1,2}|TBD|Q[1-4])[^.;\n]*",
        combined,
        flags=re.I,
    )
    return clean_text(dates[0]) if dates else ""


def infer_business_impact(project: str, body: str) -> str:
    text = f"{project} {body}".lower()
    if "cyber security" in text or "security" in text:
        return "Reduces security exposure and keeps fulfillment platforms aligned with required remediation timelines."
    if "vtm" in text and "tlvod" in text:
        return "Improves title metadata flow, reducing manual entry and improving scheduling/financial processing accuracy."
    if "wildmoka" in text or "short form" in text or "snapped fast" in text:
        return "Supports higher-volume USA Sports clipping, publishing, and distribution workflows."
    if "dataforge" in text:
        return "Enables Versant content data enrichment and production migration work."
    if "sso" in text:
        return "Protects user access controls for internal and contracted editor workflows."
    if "polaris" in text or "modernization" in text or "content hub" in text or "auto qc" in text:
        return "Advances the fulfillment modernization platform and reduces legacy workflow dependency."
    if "caption" in text or "cmp" in text:
        return "Improves caption analysis and publishing readiness for Versant workflows."
    return ""


def extract_risk(body: str, status: str) -> str:
    if status == "Completed":
        return ""
    keywords = [
        "blocked",
        "awaiting",
        "dependency",
        "dependencies",
        "not available",
        "not working",
        "need ",
        "risk",
        "permissions",
        "failed",
        "mismatch",
    ]
    sentences = re.split(r"(?<=[.!?])\s+|\n+", body)
    risk_sentences = [
        clean_text(sentence)
        for sentence in sentences
        if "no risks" not in sentence.lower() and "unblocked" not in sentence.lower() and any(key in sentence.lower() for key in keywords)
    ]
    if risk_sentences:
        return " ".join(risk_sentences[:2])
    if status in {"Blocked", "At Risk", "Monitoring"}:
        return clean_text(body)
    return ""


def leadership_ask_from_risk(risk: str) -> str:
    if not risk:
        return "No"
    lower = risk.lower()
    if any(token in lower for token in ["blocked", "awaiting", "need", "not available", "not working", "permissions", "failed"]):
        return "Yes"
    return "Review"


def extract_projects(slides: Sequence[Dict[str, Any]], deck_name: str, reporting_week: str) -> List[Dict[str, str]]:
    projects: List[Dict[str, str]] = []
    for slide in slides:
        if slide["type"] != "Project Highlights":
            continue
        raw_lines = [clean_text(line) for line in slide["lines"] if clean_text(line)]
        active = False
        i = 0
        while i < len(raw_lines):
            line = raw_lines[i]
            if line == "Status" and i + 1 < len(raw_lines) and raw_lines[i + 1] == "Project":
                active = True
                i += 2
                continue
            if not active:
                i += 1
                continue
            if line in {"Project Highlights", "Team", "Status", "Project"} or is_footer_line(line):
                i += 1
                continue
            status = normalize_status(line)
            if not status:
                i += 1
                continue
            i += 1
            title_lines: List[str] = []
            body_lines: List[str] = []
            body_started = False
            while i < len(raw_lines):
                current = raw_lines[i]
                if current == "Status" and i + 1 < len(raw_lines) and raw_lines[i + 1] == "Project":
                    break
                if normalize_status(current):
                    break
                if current in {"Project Highlights", "Team", "Status", "Project"} or is_footer_line(current):
                    i += 1
                    continue
                if body_started:
                    body_lines.append(current)
                elif not likely_body_line(current, title_lines) and should_join_title_fragment(current, title_lines):
                    title_lines.append(current)
                else:
                    body_started = True
                    body_lines.append(current)
                i += 1
            project = clean_project_title(" ".join(title_lines))
            body = clean_text(" ".join(body_lines))
            if not project or project in STATUS_OPTIONS:
                continue
            combined = f"{project} {body}"
            risk = extract_risk(body, status)
            projects.append(
                {
                    "Reporting Week": reporting_week,
                    "Workstream": infer_workstream(project),
                    "Project": project,
                    "Status": status,
                    "Owner": extract_owner(project),
                    "Target Date": extract_target_date(combined),
                    "This Week Update": body,
                    "Next Milestone": next_milestone(project, body),
                    "Business Impact": infer_business_impact(project, body),
                    "Blocker or Risk": risk,
                    "Leadership Ask": leadership_ask_from_risk(risk),
                    "Help Needed": "Clarify owner/action in weekly review." if leadership_ask_from_risk(risk) == "Yes" else "",
                    "Jira Key/Epic": "",
                    "Jira URL": "",
                    "Source Slide": str(slide["slide"]),
                    "Source Deck": deck_name,
                    "Last Updated": reporting_week,
                }
            )
    return dedupe_projects(projects)


def dedupe_projects(projects: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    seen: set[Tuple[str, str]] = set()
    deduped: List[Dict[str, str]] = []
    for row in projects:
        key = (row["Project"].lower(), row["Source Slide"])
        if key not in seen:
            seen.add(key)
            deduped.append(row)
    return deduped


def highlight_category(title: str, body: str) -> str:
    text = f"{title} {body}".lower()
    if any(token in text for token in ["golf", "nascar", "wnba", "sports", "pga", "lpga", "basketball"]):
        return "Sports/Event Volume"
    if "premiere" in text or "support sessions" in text:
        return "Operational Readiness"
    if "modernization" in text or "stakeholders" in text:
        return "Modernization"
    if "tlvod" in text or "integration" in text or "archive" in text:
        return "Workflow Efficiency"
    return "Business"


def extract_impact_volume(body: str) -> str:
    matches = re.findall(r"(?:over\s+)?\d[\d,]*(?:-\d[\d,]*)?(?:\+)?(?:\s*(?:hours|assets|games|events|clips|TB|PM ET|football games|basketball games))?", body, re.I)
    return "; ".join(clean_text(match) for match in matches[:4])


def extract_timing(body: str) -> str:
    date_terms = re.findall(
        r"(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2}(?:-\d{1,2})?|20\d{2}(?:-20\d{2})?|Sunday,?\s+[A-Za-z]+\s+\d{1,2}|Wednesday,?\s+[A-Za-z]+\s+\d{1,2}",
        body,
        re.I,
    )
    return "; ".join(clean_text(item) for item in date_terms[:4])


def extract_business_highlights(slides: Sequence[Dict[str, Any]], deck_name: str, reporting_week: str) -> List[Dict[str, str]]:
    highlights: List[Dict[str, str]] = []
    for slide in slides:
        if slide["type"] != "Business Highlights":
            continue
        lines = [clean_text(line) for line in slide["lines"] if clean_text(line) and not is_footer_line(line)]
        if "Business Highlights" in lines:
            lines = lines[lines.index("Business Highlights") + 1 :]
        current_title = ""
        current_body: List[str] = []
        for line in lines:
            if line in {"Team"}:
                continue
            if line in BUSINESS_HEADINGS:
                if current_title:
                    body = clean_text(" ".join(current_body))
                    highlights.append(
                        {
                            "Reporting Week": reporting_week,
                            "Category": highlight_category(current_title, body),
                            "Highlight": current_title,
                            "Business Context": body,
                            "Impact/Volume": extract_impact_volume(body),
                            "Timing": extract_timing(body),
                            "Leadership Note": "",
                            "Source Slide": str(slide["slide"]),
                            "Source Deck": deck_name,
                        }
                    )
                current_title = line
                current_body = []
            elif current_title:
                current_body.append(line)
        if current_title:
            body = clean_text(" ".join(current_body))
            highlights.append(
                {
                    "Reporting Week": reporting_week,
                    "Category": highlight_category(current_title, body),
                    "Highlight": current_title,
                    "Business Context": body,
                    "Impact/Volume": extract_impact_volume(body),
                    "Timing": extract_timing(body),
                    "Leadership Note": "",
                    "Source Slide": str(slide["slide"]),
                    "Source Deck": deck_name,
                }
            )
    return highlights


def add_metric(
    rows: List[Dict[str, str]],
    period: str,
    area: str,
    metric: str,
    value: str,
    unit: str,
    context: str,
    slide: Any,
    deck_name: str,
) -> None:
    rows.append(
        {
            "Reporting Period": period,
            "Metric Area": area,
            "Metric": metric,
            "Value": value,
            "Unit/Breakdown": unit,
            "Leadership Context": context,
            "Source Slide": str(slide),
            "Source Deck": deck_name,
        }
    )


def extract_department_metrics(slides: Sequence[Dict[str, Any]], deck_name: str) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for slide in slides:
        if slide["type"] != "Department Highlights":
            continue
        text = slide["text"]
        period = "March" if "March Highlights" in text else "December" if "December" in text else ""
        if "Net New Deliveries to Amagi" in text:
            match = re.search(r"Linear\s+([\d,]+)?\s*Net New Deliveries to Amagi", text)
            if match and match.group(1):
                add_metric(rows, period, "Linear", "Net New Deliveries to Amagi", match.group(1), "deliveries", "Topline linear delivery volume.", slide["slide"], deck_name)
            for media, pattern in [
                ("Longform", r"% Longform \(([\d,]+)\)"),
                ("Shortform", r"% Shortform: \(([\d,]+)\)"),
                ("Graphics", r"% Graphics: \(([\d,]+)\)"),
            ]:
                found = re.search(pattern, text)
                if found:
                    add_metric(rows, period, "Linear", f"{media} registrations", found.group(1), "assets", "Media-type mix for linear registrations.", slide["slide"], deck_name)
        found = re.search(r"([\d,]+) Assets\s*/\s*([\d.]+ TB)\s*transferred from NBCU to Versant", text)
        if found:
            add_metric(rows, period, "Linear", "NBCU to Versant transfer count", found.group(1), "assets", "Volume moved into Versant workflows.", slide["slide"], deck_name)
            add_metric(rows, period, "Linear", "NBCU to Versant transfer size", found.group(2), "data volume", "Storage/network footprint for transferred assets.", slide["slide"], deck_name)
        found = re.search(r"iComm\s+([\d,]+)\s+iComm deliveries", text, re.I)
        if found:
            add_metric(rows, period, "iComm", "iComm deliveries", found.group(1), "deliveries", "Downstream iComm delivery volume.", slide["slide"], deck_name)
        for label, pattern in [("TVN | Akamai", r"TVN \| Akamai: ([\d,]+)"), ("STB VOD DAI | CMC", r"STB VOD DAI \| CMC: ([\d,]+)")]:
            found = re.search(pattern, text)
            if found:
                add_metric(rows, period, "iComm", label, found.group(1), "deliveries", "iComm delivery channel breakdown.", slide["slide"], deck_name)
        found = re.search(r"TVE/VOD\s+([\d,]+)\s+Longform Deliveries through FP\s+([\d,]+)\s+Titles Delivered", text)
        if found:
            add_metric(rows, period, "TVE/VOD", "Longform deliveries through FP", found.group(1), "deliveries", "Fulfillment Pipeline longform throughput.", slide["slide"], deck_name)
            add_metric(rows, period, "TVE/VOD", "Titles delivered", found.group(2), "titles", "Title delivery volume.", slide["slide"], deck_name)
        found = re.search(r"Digital\s+([\d,]+)\s+Shortform clips ingested and published", text)
        if found:
            add_metric(rows, period, "Digital", "Shortform clips ingested and published", found.group(1), "clips", "Digital publishing throughput.", slide["slide"], deck_name)
        for network, pattern in [
            ("CNBC", r"CNBC: ([\d,]+) -> ([\d.]+%)"),
            ("E!", r"E!: ([\d,]+) -> ([\d.]+%)"),
            ("Golf Channel", r"Golf Channel: ([\d,]+) -> ([\d.]+%)"),
            ("MSNOW", r"MSNOW: ([\d,]+) -> ([\d.]+%)"),
            ("Oxygen", r"Oxygen: ([\d,]+) -> ([\d.]+%)"),
            ("Syfy", r"Syfy: ([\d,]+) -> ([\d.]+%)"),
            ("USA Network", r"USA Network: ([\d,]+) -> ([\d.]+%)"),
        ]:
            found = re.search(pattern, text)
            if found:
                add_metric(rows, period, "Digital", f"{network} shortform clips", found.group(1), found.group(2), "Brand mix for shortform output.", slide["slide"], deck_name)
        found = re.search(r"total number of asset registrations.*?December is ([\d,]+)", text, re.I)
        if found:
            add_metric(rows, "December", "Registrations", "Total asset registrations", found.group(1), "registrations", "Appendix benchmark from registration charts.", slide["slide"], deck_name)
        for media, pattern in [
            ("Longform share", r"Longform\s*=?\s*([\d.]+%)"),
            ("Shortform share", r"Shortform\s*=?\s*([\d.]+%)"),
            ("Graphics share", r"Graphics\s*=?\s*([\d.]+%)"),
        ]:
            found = re.search(pattern, text)
            if found:
                add_metric(rows, "December", "Registrations", media, found.group(1), "share", "Appendix media-type share.", slide["slide"], deck_name)
        found = re.search(r"Approximately ~?([\dKk,]+) individual asset files were received", text)
        if found:
            add_metric(rows, "December", "DIVA Transfer", "Individual asset files received", found.group(1), "files", "DIVA transfer operating volume.", slide["slide"], deck_name)
        found = re.search(r"roughly ~?([\d.]+ TB) of data", text)
        if found:
            add_metric(rows, "December", "DIVA Transfer", "Total data volume", found.group(1), "data volume", "DIVA storage/network footprint.", slide["slide"], deck_name)
        found = re.search(r"total number of iCOMM registrations is ([\d,]+)", text, re.I)
        if found:
            add_metric(rows, "December", "iComm", "iCOMM registrations", found.group(1), "registrations", "Appendix iComm benchmark.", slide["slide"], deck_name)
        for label, pattern in [("Akamai", r"Akamai: ([\d,]+)"), ("CMC", r"CMC: ([\d,]+)")]:
            found = re.search(pattern, text)
            if found:
                add_metric(rows, "December", "iComm", label, found.group(1), "registrations", "iComm registration channel breakdown.", slide["slide"], deck_name)
    return dedupe_metrics(rows)


def dedupe_metrics(rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    deduped: List[Dict[str, str]] = []
    index_by_key: Dict[Tuple[str, str, str, str], int] = {}
    for row in rows:
        key = (
            row.get("Reporting Period", ""),
            row.get("Metric Area", ""),
            row.get("Metric", ""),
            row.get("Unit/Breakdown", ""),
        )
        if key not in index_by_key:
            index_by_key[key] = len(deduped)
            deduped.append(dict(row))
            continue
        existing = deduped[index_by_key[key]]
        if existing.get("Value") == row.get("Value"):
            continue
        if not existing.get("Value") and row.get("Value"):
            deduped[index_by_key[key]] = dict(row)
            continue
        deduped.append(dict(row))
    return deduped


def build_blockers(projects: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for project in projects:
        if project.get("Blocker or Risk") or project.get("Status") in {"Blocked", "At Risk", "Monitoring"}:
            rows.append(
                {
                    "Reporting Week": project.get("Reporting Week", ""),
                    "Workstream": project.get("Workstream", ""),
                    "Project": project.get("Project", ""),
                    "Status": project.get("Status", ""),
                    "Blocker or Risk": project.get("Blocker or Risk", ""),
                    "Impact": project.get("Business Impact", ""),
                    "Owner": project.get("Owner", ""),
                    "Help Needed": project.get("Help Needed", ""),
                    "Target Date": project.get("Target Date", ""),
                    "Source Slide": project.get("Source Slide", ""),
                }
            )
    return rows


def jira_auth_header(email: str, token: str) -> str:
    encoded = base64.b64encode(f"{email}:{token}".encode("utf-8")).decode("ascii")
    return f"Basic {encoded}"


def jira_get_json(base_url: str, path: str, params: Dict[str, Any], auth_header: str) -> Dict[str, Any]:
    url = base_url.rstrip("/") + path + "?" + urllib.parse.urlencode(params, doseq=True)
    request = urllib.request.Request(url, headers={"Accept": "application/json", "Authorization": auth_header})
    try:
        with urllib.request.urlopen(request, timeout=45) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Jira request failed: HTTP {exc.code} {body[:500]}") from exc


def issue_field(fields: Dict[str, Any], name: str, default: str = "") -> str:
    value = fields.get(name)
    if value is None:
        return default
    if isinstance(value, dict):
        return str(value.get("name") or value.get("displayName") or value.get("key") or value.get("value") or default)
    if isinstance(value, list):
        values = []
        for item in value:
            if isinstance(item, dict):
                values.append(str(item.get("name") or item.get("value") or item.get("key") or ""))
            else:
                values.append(str(item))
        return ", ".join(item for item in values if item)
    return str(value)


def jira_issue_to_row(issue: Dict[str, Any], base_url: str, source: str, query_name: str, run_id: str) -> Dict[str, str]:
    fields = issue.get("fields", {})
    status_obj = fields.get("status") or {}
    status_category = status_obj.get("statusCategory", {}) if isinstance(status_obj, dict) else {}
    sprint_value = issue_field(fields, "customfield_10020")
    epic_key = issue_field(fields, "customfield_10014")
    epic_name = issue_field(fields, "customfield_10011")
    story_points = issue_field(fields, "customfield_10016")
    return {
        "Run ID": run_id,
        "Source": source,
        "Query/Board": query_name,
        "Issue Key": issue.get("key", ""),
        "Issue Type": issue_field(fields, "issuetype"),
        "Summary": issue_field(fields, "summary"),
        "Status": issue_field(fields, "status"),
        "Status Category": str(status_category.get("name", "")),
        "Assignee": issue_field(fields, "assignee", "Unassigned"),
        "Priority": issue_field(fields, "priority"),
        "Labels": issue_field(fields, "labels"),
        "Components": issue_field(fields, "components"),
        "Fix Versions": issue_field(fields, "fixVersions"),
        "Due Date": issue_field(fields, "duedate"),
        "Updated": issue_field(fields, "updated"),
        "Sprint": sprint_value,
        "Epic Key": epic_key,
        "Epic Name": epic_name,
        "Story Points": story_points,
        "URL": f"{base_url.rstrip()}/browse/{issue.get('key', '')}" if issue.get("key") else "",
    }


def fetch_jira(config_path: Path) -> List[Dict[str, str]]:
    config = json.loads(config_path.read_text(encoding="utf-8"))
    base_url = config["base_url"]
    email = os.getenv(config.get("email_env", "JIRA_EMAIL"), "")
    token = os.getenv(config.get("api_token_env", "JIRA_API_TOKEN"), "")
    if not email or not token:
        raise RuntimeError("Set Jira credentials in the configured email/token environment variables before fetching Jira data.")
    auth_header = jira_auth_header(email, token)
    run_id = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    fields = config.get(
        "fields",
        [
            "summary",
            "status",
            "assignee",
            "priority",
            "labels",
            "components",
            "fixVersions",
            "duedate",
            "updated",
            "issuetype",
            "customfield_10020",
            "customfield_10014",
            "customfield_10011",
            "customfield_10016",
        ],
    )
    rows: List[Dict[str, str]] = []
    for query in config.get("queries", []):
        token_value = ""
        while True:
            params: Dict[str, Any] = {
                "jql": query["jql"],
                "maxResults": query.get("max_results", 100),
                "fields": ",".join(fields),
            }
            if token_value:
                params["nextPageToken"] = token_value
            data = jira_get_json(base_url, "/rest/api/3/search/jql", params, auth_header)
            for issue in data.get("issues", []):
                rows.append(jira_issue_to_row(issue, base_url, "JQL", query.get("name", query["jql"]), run_id))
            token_value = data.get("nextPageToken") or ""
            if not token_value:
                break
    for board in config.get("boards", []):
        start_at = 0
        max_results = board.get("max_results", 100)
        while True:
            params = {
                "startAt": start_at,
                "maxResults": max_results,
                "fields": ",".join(fields),
            }
            if board.get("jql"):
                params["jql"] = board["jql"]
            data = jira_get_json(base_url, f"/rest/agile/1.0/board/{board['board_id']}/issue", params, auth_header)
            for issue in data.get("issues", []):
                rows.append(jira_issue_to_row(issue, base_url, "Board", board.get("name", str(board["board_id"])), run_id))
            start_at += len(data.get("issues", []))
            if start_at >= int(data.get("total", start_at)):
                break
    return rows


def jira_status_to_tracker_status(status: str, status_category: str) -> str:
    status_lower = clean_text(status).lower()
    category_lower = clean_text(status_category).lower()
    if "block" in status_lower or "impediment" in status_lower:
        return "Blocked"
    if category_lower == "done" or status_lower in {"done", "closed", "resolved", "complete", "completed"}:
        return "Completed"
    if any(token in status_lower for token in ["risk", "at risk"]):
        return "At Risk"
    if any(token in status_lower for token in ["monitor", "review", "waiting", "hold", "dependency"]):
        return "Monitoring"
    if category_lower in {"in progress", "indeterminate"} or any(token in status_lower for token in ["progress", "active", "doing"]):
        return "On Track"
    return "Upcoming"


def infer_workstream_from_jira(row: Dict[str, str]) -> str:
    text = " ".join(
        [
            row.get("Components", ""),
            row.get("Labels", ""),
            row.get("Query/Board", ""),
            row.get("Summary", ""),
            row.get("Epic Name", ""),
        ]
    ).lower()
    if any(token in text for token in ["polaris", "modernization", "content hub", "auto qc", "sdvi"]):
        return "Fulfillment Modernization"
    if any(token in text for token in ["cvp", "wildmoka", "viewlift"]):
        return "CVP"
    if any(token in text for token in ["tlvod", "fp", "fulfillment pipeline", "vtm"]):
        return "TLVOD/Fulfillment Pipeline"
    if any(token in text for token in ["otto", "media broker", "dataforge", "rundown"]):
        return "OTTO/Media Broker"
    return "Business"


def jira_row_to_project(row: Dict[str, str], reporting_week: str) -> Dict[str, str]:
    issue_key = row.get("Issue Key", "")
    summary = row.get("Summary", "")
    status = jira_status_to_tracker_status(row.get("Status", ""), row.get("Status Category", ""))
    project_name = f"{issue_key} - {summary}" if issue_key else summary
    risk = ""
    priority = row.get("Priority", "")
    if status in {"Blocked", "At Risk", "Monitoring"}:
        risk = f"Jira status is {row.get('Status', status)}."
    elif priority.lower() in {"highest", "critical", "blocker"}:
        risk = f"High-priority Jira item: {priority}."
    milestone_parts = [row.get("Sprint", ""), row.get("Fix Versions", "")]
    return {
        "Reporting Week": reporting_week,
        "Workstream": infer_workstream_from_jira(row),
        "Project": clean_text(project_name) or "Jira issue",
        "Status": status,
        "Owner": row.get("Assignee", "") or "Unassigned",
        "Target Date": row.get("Due Date", ""),
        "This Week Update": clean_text(" | ".join(part for part in [row.get("Issue Type", ""), row.get("Status", ""), priority] if part)),
        "Next Milestone": clean_text(" | ".join(part for part in milestone_parts if part)),
        "Business Impact": clean_text("Components: " + row.get("Components", "")) if row.get("Components") else "",
        "Blocker or Risk": risk,
        "Leadership Ask": leadership_ask_from_risk(risk),
        "Help Needed": "Review Jira blocker/dependency." if status in {"Blocked", "At Risk"} else "",
        "Jira Key/Epic": issue_key or row.get("Epic Key", ""),
        "Jira URL": row.get("URL", ""),
        "Source Slide": "Jira",
        "Source Deck": row.get("Query/Board", "Jira"),
        "Last Updated": row.get("Updated", "")[:10] or reporting_week,
    }


def jira_rows_to_projects(rows: Sequence[Dict[str, str]], reporting_week: str) -> List[Dict[str, str]]:
    return [jira_row_to_project(row, reporting_week) for row in rows if row.get("Issue Key") or row.get("Summary")]


def write_csv(path: Path, headers: Sequence[str], rows: Sequence[Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def ensure_rows(rows: Sequence[Dict[str, str]], headers: Sequence[str]) -> List[Dict[str, str]]:
    if rows:
        return list(rows)
    return [{header: "" for header in headers}]


def add_table(ws, name: str, headers: Sequence[str], rows: Sequence[Dict[str, str]], style: str = "TableStyleMedium2") -> None:
    ws.append(list(headers))
    for row in ensure_rows(rows, headers):
        ws.append([row.get(header, "") for header in headers])
    ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ref


def style_sheet(ws, widths: Optional[Dict[str, int]] = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if widths:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
    else:
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 18


def add_dropdown(ws, column_name: str, headers: Sequence[str], values: Sequence[str], max_rows: int = 500) -> None:
    try:
        column_idx = headers.index(column_name) + 1
    except ValueError:
        return
    column_letter = get_column_letter(column_idx)
    validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{max_rows}")


def populate_summary(ws, projects: Sequence[Dict[str, str]], blockers: Sequence[Dict[str, str]], highlights: Sequence[Dict[str, str]], metrics: Sequence[Dict[str, str]], deck_name: str, reporting_week: str) -> None:
    ws.title = "Leadership Summary"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Weekly Fulfillment Leadership Update"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws["A2"] = f"Reporting week: {reporting_week}"
    ws["A3"] = f"Source: {deck_name}"

    cards = [
        ("Total Projects", len(projects)),
        ("On Track", sum(1 for row in projects if row.get("Status") == "On Track")),
        ("Completed", sum(1 for row in projects if row.get("Status") == "Completed")),
        ("Needs Leadership", sum(1 for row in projects if row.get("Leadership Ask") == "Yes")),
        ("Blockers/Risks", len(blockers)),
        ("Business Highlights", len(highlights)),
    ]
    start_col = 1
    for idx, (label, value) in enumerate(cards):
        col = start_col + idx * 2
        ws.cell(row=5, column=col, value=label)
        ws.cell(row=6, column=col, value=value)
        ws.cell(row=5, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=5, column=col).fill = PatternFill("solid", fgColor="1F4E78")
        ws.cell(row=6, column=col).font = Font(size=16, bold=True)
        ws.cell(row=6, column=col).fill = PatternFill("solid", fgColor="D9EAF7")
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws["A9"] = "Leadership Attention"
    ws["A9"].font = Font(size=14, bold=True, color="1F4E78")
    row_num = 10
    if blockers:
        for blocker in blockers[:8]:
            ws.cell(row=row_num, column=1, value=blocker.get("Project"))
            ws.cell(row=row_num, column=2, value=blocker.get("Status"))
            ws.cell(row=row_num, column=3, value=blocker.get("Blocker or Risk"))
            ws.cell(row=row_num, column=4, value=blocker.get("Help Needed"))
            row_num += 1
    else:
        ws.cell(row=row_num, column=1, value="No blockers or at-risk items were found in this refresh.")
        row_num += 1

    ws[f"A{row_num + 1}"] = "Business Highlights"
    ws[f"A{row_num + 1}"].font = Font(size=14, bold=True, color="1F4E78")
    row_num += 2
    for highlight in highlights[:8]:
        ws.cell(row=row_num, column=1, value=highlight.get("Highlight"))
        ws.cell(row=row_num, column=2, value=highlight.get("Category"))
        ws.cell(row=row_num, column=3, value=highlight.get("Impact/Volume"))
        ws.cell(row=row_num, column=4, value=highlight.get("Business Context"))
        row_num += 1

    ws[f"A{row_num + 1}"] = "Key Metrics"
    ws[f"A{row_num + 1}"].font = Font(size=14, bold=True, color="1F4E78")
    row_num += 2
    for metric in metrics[:10]:
        ws.cell(row=row_num, column=1, value=metric.get("Metric Area"))
        ws.cell(row=row_num, column=2, value=metric.get("Metric"))
        ws.cell(row=row_num, column=3, value=metric.get("Value"))
        ws.cell(row=row_num, column=4, value=metric.get("Unit/Breakdown"))
        row_num += 1

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in {"A": 34, "B": 18, "C": 46, "D": 60, "E": 18, "F": 18, "G": 18, "H": 18, "I": 18, "J": 18, "K": 18}.items():
        ws.column_dimensions[col].width = width


def populate_config(ws, deck_path: Path, reporting_week: str) -> None:
    rows = [
        {"Setting": "Reporting Week", "Value": reporting_week, "Notes": "Used to stamp new rows during refresh."},
        {"Setting": "Source Deck", "Value": str(deck_path), "Notes": "Run refresh_tracker.py with --deck to update this."},
        {"Setting": "Jira Config", "Value": "config/jira.json", "Notes": "Copy config/jira.example.json to config/jira.json and edit for your boards/JQL."},
        {"Setting": "Jira Credentials", "Value": "JIRA_EMAIL / JIRA_API_TOKEN", "Notes": "Keep secrets in environment variables, not in the config file."},
        {"Setting": "Status Values", "Value": ", ".join(STATUS_OPTIONS), "Notes": "Used by Project Tracker validation."},
        {"Setting": "Workstreams", "Value": ", ".join(WORKSTREAM_OPTIONS), "Notes": "Used by Project Tracker validation."},
    ]
    add_table(ws, "ConfigTable", ["Setting", "Value", "Notes"], rows, style="TableStyleMedium4")
    style_sheet(ws, {"A": 24, "B": 48, "C": 78})


def create_weekly_intake_rows(reporting_week: str) -> List[Dict[str, str]]:
    return [
        {
            "Reporting Week": reporting_week,
            "Workstream": "",
            "Project": "",
            "Status": "",
            "Owner": "",
            "Target Date": "",
            "This Week Update": "",
            "Next Milestone": "",
            "Business Impact": "",
            "Blocker or Risk": "",
            "Leadership Ask": "",
            "Help Needed": "",
            "Jira Key/Epic": "",
            "Jira URL": "",
            "Source Slide": "Manual",
            "Source Deck": "",
            "Last Updated": reporting_week,
        }
        for _ in range(15)
    ]


def write_workbook(
    path: Path,
    deck_path: Path,
    projects: Sequence[Dict[str, str]],
    blockers: Sequence[Dict[str, str]],
    highlights: Sequence[Dict[str, str]],
    metrics: Sequence[Dict[str, str]],
    jira_rows: Sequence[Dict[str, str]],
    source_rows: Sequence[Dict[str, str]],
    reporting_week: str,
) -> None:
    if Workbook is None:
        raise RuntimeError("Install openpyxl to export the optional workbook. The browser tracker refresh does not require it.")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    populate_summary(summary, projects, blockers, highlights, metrics, deck_path.name, reporting_week)

    project_ws = wb.create_sheet("Project Tracker")
    add_table(project_ws, "ProjectTracker", PROJECT_HEADERS, projects, style="TableStyleMedium2")
    add_dropdown(project_ws, "Status", PROJECT_HEADERS, STATUS_OPTIONS)
    add_dropdown(project_ws, "Workstream", PROJECT_HEADERS, WORKSTREAM_OPTIONS)
    add_dropdown(project_ws, "Leadership Ask", PROJECT_HEADERS, ["No", "Yes", "Review"])
    style_sheet(
        project_ws,
        {"A": 16, "B": 24, "C": 42, "D": 14, "E": 18, "F": 16, "G": 58, "H": 34, "I": 44, "J": 44, "K": 16, "L": 36, "M": 18, "N": 28, "O": 12, "P": 34, "Q": 16},
    )

    blocker_ws = wb.create_sheet("Blockers and Risks")
    add_table(blocker_ws, "BlockerRisk", BLOCKER_HEADERS, blockers, style="TableStyleMedium3")
    add_dropdown(blocker_ws, "Status", BLOCKER_HEADERS, STATUS_OPTIONS)
    style_sheet(blocker_ws, {"A": 16, "B": 24, "C": 42, "D": 14, "E": 58, "F": 44, "G": 18, "H": 38, "I": 16, "J": 12})

    highlight_ws = wb.create_sheet("Business Highlights")
    add_table(highlight_ws, "BusinessHighlights", HIGHLIGHT_HEADERS, highlights, style="TableStyleMedium6")
    style_sheet(highlight_ws, {"A": 16, "B": 22, "C": 34, "D": 74, "E": 26, "F": 28, "G": 44, "H": 12, "I": 34})

    metric_ws = wb.create_sheet("Department Metrics")
    add_table(metric_ws, "DepartmentMetrics", METRIC_HEADERS, metrics, style="TableStyleMedium7")
    style_sheet(metric_ws, {"A": 18, "B": 20, "C": 42, "D": 18, "E": 26, "F": 54, "G": 12, "H": 34})

    jira_ws = wb.create_sheet("Jira Raw")
    add_table(jira_ws, "JiraRaw", JIRA_HEADERS, jira_rows, style="TableStyleMedium5")
    style_sheet(jira_ws, {"A": 18, "B": 12, "C": 24, "D": 14, "E": 18, "F": 60, "G": 18, "H": 18, "I": 24, "J": 16, "K": 24, "L": 24, "M": 24, "N": 16, "O": 22, "P": 26, "Q": 18, "R": 24, "S": 16, "T": 42})

    intake_ws = wb.create_sheet("Weekly Intake")
    add_table(intake_ws, "WeeklyIntake", PROJECT_HEADERS, create_weekly_intake_rows(reporting_week), style="TableStyleMedium9")
    add_dropdown(intake_ws, "Status", PROJECT_HEADERS, STATUS_OPTIONS)
    add_dropdown(intake_ws, "Workstream", PROJECT_HEADERS, WORKSTREAM_OPTIONS)
    add_dropdown(intake_ws, "Leadership Ask", PROJECT_HEADERS, ["No", "Yes", "Review"])
    style_sheet(
        intake_ws,
        {"A": 16, "B": 24, "C": 42, "D": 14, "E": 18, "F": 16, "G": 58, "H": 34, "I": 44, "J": 44, "K": 16, "L": 36, "M": 18, "N": 28, "O": 12, "P": 34, "Q": 16},
    )

    source_ws = wb.create_sheet("Source Slides")
    add_table(source_ws, "SourceSlides", SOURCE_HEADERS, source_rows, style="TableStyleMedium1")
    style_sheet(source_ws, {"A": 10, "B": 22, "C": 100, "D": 34})

    config_ws = wb.create_sheet("Config")
    populate_config(config_ws, deck_path, reporting_week)

    wb.save(path)


def write_leadership_brief(
    path: Path,
    projects: Sequence[Dict[str, str]],
    blockers: Sequence[Dict[str, str]],
    highlights: Sequence[Dict[str, str]],
    metrics: Sequence[Dict[str, str]],
    deck_path: Path,
    reporting_week: str,
) -> None:
    lines = [
        f"# Weekly Fulfillment Leadership Update - {reporting_week}",
        "",
        f"Source: {deck_path.name}",
        "",
        "## Snapshot",
        "",
        f"- Total projects tracked: {len(projects)}",
        f"- On track: {sum(1 for row in projects if row.get('Status') == 'On Track')}",
        f"- Completed: {sum(1 for row in projects if row.get('Status') == 'Completed')}",
        f"- Blockers/risks: {len(blockers)}",
        f"- Business highlights: {len(highlights)}",
        "",
        "## Leadership Attention",
        "",
    ]
    if blockers:
        for blocker in blockers[:8]:
            lines.append(f"- {blocker.get('Project')} [{blocker.get('Status')}]: {blocker.get('Blocker or Risk') or 'Review status and next action.'}")
    else:
        lines.append("- No blockers or at-risk items were found in this refresh.")
    lines.extend(["", "## Business Highlights", ""])
    for highlight in highlights[:8]:
        impact = f" ({highlight.get('Impact/Volume')})" if highlight.get("Impact/Volume") else ""
        lines.append(f"- {highlight.get('Highlight')}{impact}: {highlight.get('Business Context')}")
    lines.extend(["", "## Key Metrics", ""])
    for metric in metrics[:10]:
        value = " ".join(part for part in [metric.get("Value"), metric.get("Unit/Breakdown")] if part)
        lines.append(f"- {metric.get('Metric Area')} - {metric.get('Metric')}: {value}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def write_dashboard_data(
    path: Path,
    projects: Sequence[Dict[str, str]],
    blockers: Sequence[Dict[str, str]],
    highlights: Sequence[Dict[str, str]],
    metrics: Sequence[Dict[str, str]],
    jira_rows: Sequence[Dict[str, str]],
    source_rows: Sequence[Dict[str, str]],
    deck_path: Path,
    reporting_week: str,
) -> None:
    payload = {
        "generatedAt": dt.datetime.now().isoformat(timespec="seconds"),
        "reportingWeek": reporting_week,
        "deckName": deck_path.name,
        "projects": list(projects),
        "blockers": list(blockers),
        "highlights": list(highlights),
        "metrics": list(metrics),
        "jiraRows": list(jira_rows),
        "sourceRows": list(source_rows),
        "briefPath": "outputs/leadership_brief.md",
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "window.TRACKER_DATA = "
        + json.dumps(payload, indent=2, ensure_ascii=True)
        + ";\n",
        encoding="utf-8",
    )


def source_slide_rows(slides: Sequence[Dict[str, Any]], deck_name: str) -> List[Dict[str, str]]:
    return [
        {
            "Slide": str(slide["slide"]),
            "Slide Type": slide["type"],
            "Extracted Text": slide["text"],
            "Source Deck": deck_name,
        }
        for slide in slides
    ]


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Refresh weekly leadership tracker from a PPTX deck and optional Jira export.")
    parser.add_argument("--deck", type=Path, help="Optional path to the weekly PowerPoint deck.")
    parser.add_argument("--reporting-week", default=dt.date.today().isoformat(), help="Reporting week/date to stamp into tracker rows.")
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parents[1], help="Tracker package directory.")
    parser.add_argument("--fetch-jira", action="store_true", help="Fetch Jira rows using --jira-config.")
    parser.add_argument("--jira-config", type=Path, default=None, help="Path to Jira config JSON. Required with --fetch-jira.")
    parser.add_argument("--jira-csv", type=Path, default=None, help="Optional existing Jira CSV to include without fetching.")
    parser.add_argument("--jira-as-projects", action="store_true", help="Add fetched/imported Jira issues to the main Project Tracker rows.")
    return parser.parse_args(argv)


def read_jira_csv(path: Path) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    output_dir = args.output_dir.resolve()
    data_dir = output_dir / "data"
    outputs_dir = output_dir / "outputs"
    deck_path = args.deck.expanduser().resolve() if args.deck else None
    if deck_path and not deck_path.exists():
        print(f"Deck not found: {deck_path}", file=sys.stderr)
        return 2

    if not deck_path and not args.fetch_jira and not args.jira_csv:
        print("Provide --deck, --fetch-jira, or --jira-csv so the tracker has a source to refresh from.", file=sys.stderr)
        return 2

    source_path = deck_path or Path("Jira connected tracker")
    slides: List[Dict[str, Any]] = []
    projects: List[Dict[str, str]] = []
    highlights: List[Dict[str, str]] = []
    metrics: List[Dict[str, str]] = []
    source_rows: List[Dict[str, str]] = []

    if deck_path:
        slides = extract_slides(deck_path)
        projects = extract_projects(slides, deck_path.name, args.reporting_week)
        highlights = extract_business_highlights(slides, deck_path.name, args.reporting_week)
        metrics = extract_department_metrics(slides, deck_path.name)
        source_rows = source_slide_rows(slides, deck_path.name)

    jira_rows: List[Dict[str, str]] = []
    if args.fetch_jira:
        if not args.jira_config:
            print("--jira-config is required with --fetch-jira", file=sys.stderr)
            return 2
        jira_rows = fetch_jira(args.jira_config.expanduser().resolve())
    elif args.jira_csv:
        jira_rows = read_jira_csv(args.jira_csv.expanduser().resolve())

    jira_project_rows: List[Dict[str, str]] = []
    if args.jira_as_projects and jira_rows:
        jira_project_rows = jira_rows_to_projects(jira_rows, args.reporting_week)
        projects = dedupe_projects([*projects, *jira_project_rows])

    blockers = build_blockers(projects)

    write_csv(data_dir / "projects.csv", PROJECT_HEADERS, projects)
    write_csv(data_dir / "blockers_and_risks.csv", BLOCKER_HEADERS, blockers)
    write_csv(data_dir / "business_highlights.csv", HIGHLIGHT_HEADERS, highlights)
    write_csv(data_dir / "department_metrics.csv", METRIC_HEADERS, metrics)
    write_csv(data_dir / "jira_raw.csv", JIRA_HEADERS, jira_rows)
    write_csv(data_dir / "source_slides.csv", SOURCE_HEADERS, source_rows)

    brief_path = outputs_dir / "leadership_brief.md"
    dashboard_data_path = data_dir / "tracker-data.js"
    write_leadership_brief(brief_path, projects, blockers, highlights, metrics, source_path, args.reporting_week)
    write_dashboard_data(dashboard_data_path, projects, blockers, highlights, metrics, jira_rows, source_rows, source_path, args.reporting_week)

    print(f"Created leadership brief: {brief_path}")
    print(f"Created dashboard data: {dashboard_data_path}")
    print(f"Extracted {len(projects)} projects, {len(blockers)} blockers/risks, {len(highlights)} business highlights, {len(metrics)} metrics.")
    if not jira_rows:
        print("No Jira rows included yet. Configure config/jira.json and rerun with --fetch-jira when ready.")
    else:
        print(f"Included {len(jira_rows)} Jira rows.")
        if args.jira_as_projects:
            print(f"Added {len(jira_project_rows)} Jira rows to the main Project Tracker.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
